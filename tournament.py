from typing import Dict, Callable, Tuple, List
from collections import defaultdict
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import matplotlib.pyplot as plt
import numpy as np

from engine import IteratedPrisonersDilemma

# Import strategies explicitly
from strategies import (
    clara,
    victor,
    miles,
    elena,
    isabella,
    nathan,
    gabriel,
    iris,
    lucas,
    samuel,
    emily
)


Move = str
History = List[Move]


class Tournament:
    def __init__(
        self,
        strategies: Dict[str, Callable] = None,
        rounds: int = 200,
        noise: float = 0.0,
        seed: int = None,
    ):
        if strategies is None:
            strategies = {
                "Clara": clara,
                "Victor": victor,
                "Miles": miles,
                "Elena": elena,
                "Isabella": isabella,
                "Nathan": nathan,
                "Gabriel": gabriel,
                "Iris": iris,
                "Lucas": lucas,
                "Samuel": samuel,
                "Emily": emily
            }

        self.strategies = strategies
        self.game = IteratedPrisonersDilemma(
            rounds=rounds,
            noise=noise,
            seed=seed,
        )
        self.results = {}
        self.match_data = []

    def run_round_robin(self):
        total_scores = defaultdict(int)
        match_results = []

        names = list(self.strategies.keys())

        # Self-play
        for name in names:
            strat = self.strategies[name]
            score1, score2, history1, history2 = self.game.play_match(strat, strat)

            total_scores[name] += score1  # score1 == score2

            match_results.append({
                "player1": name,
                "player2": name,
                "score1": score1,
                "score2": score2,
                "history1": history1,
                "history2": history2,
            })

        # Unique pair matches only
        for name1, name2 in combinations(names, 2):
            strat1 = self.strategies[name1]
            strat2 = self.strategies[name2]

            score1, score2, history1, history2 = self.game.play_match(
                strat1,
                strat2,
            )

            total_scores[name1] += score1
            total_scores[name2] += score2

            match_results.append({
                "player1": name1,
                "player2": name2,
                "score1": score1,
                "score2": score2,
                "history1": history1,
                "history2": history2,
            })

        self.results = dict(total_scores)
        self.match_data = match_results

        return self.results, self.match_data

    def ranked_results(self):
        if not self.results:
            self.run_round_robin()

        return sorted(
            self.results.items(),
            key=lambda x: x[1],
            reverse=True,
        )

    def export_match_to_excel(self, history1, history2, player1, player2, filename="match_output.xlsx"):
        """
        Exports match timeline to Excel (vertical layout).
        Column A: Round number (starting from 1)
        Column B: Player 1 move (colored)
        Column C: Player 2 move (colored)
        """

        wb = Workbook()
        ws = wb.active
        ws.title = f"{player1}_vs_{player2}"

        rounds = len(history1)

        green_fill = PatternFill(start_color="00C853", end_color="00C853", fill_type="solid")
        red_fill = PatternFill(start_color="D50000", end_color="D50000", fill_type="solid")

        # Headers
        ws.cell(row=1, column=1).value = "Round"
        ws.cell(row=1, column=2).value = player1
        ws.cell(row=1, column=3).value = player2

        # Fill rows
        for i in range(rounds):
            row_index = i + 2  # data starts at row 2

            ws.cell(row=row_index, column=1).value = i + 1  # rounds start at 1

            cell1 = ws.cell(row=row_index, column=2)
            cell2 = ws.cell(row=row_index, column=3)

            cell1.value = ""
            cell2.value = ""

            cell1.fill = green_fill if history1[i] == "C" else red_fill
            cell2.fill = green_fill if history2[i] == "C" else red_fill

            cell1.alignment = Alignment(horizontal="center")
            cell2.alignment = Alignment(horizontal="center")

        # Clean column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 5

        wb.save(filename)
