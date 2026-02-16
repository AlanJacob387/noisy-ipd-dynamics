from tournament import Tournament
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment


def main():
    ROUNDS = 10000
    NOISE = 0.2
    SEED = 314232

    print("Prisoner's Dilemma Tournament")
    print(f"Rounds per match: {ROUNDS}")
    print(f"Noise level: {NOISE}")
    print("-" * 50)

    tournament = Tournament(
        rounds=ROUNDS,
        noise=NOISE,
        seed=SEED,
    )

    results, match_data = tournament.run_round_robin()

    # --- Print ranking by average per round ---
    num_strategies = len(tournament.strategies)
    total_rounds_per_strategy = num_strategies * ROUNDS

    averages = {
        name: score / total_rounds_per_strategy
        for name, score in results.items()
    }

    ranked = sorted(averages.items(), key=lambda x: x[1], reverse=True)

    print("Ranking by Average Payoff per Round")
    print("-" * 50)
    for rank, (name, avg) in enumerate(ranked, start=1):
        print(f"{rank}. {name} | Avg per round: {avg:.3f}")
    print("-" * 50)

    # --- Create Excel workbook for ALL matches ---
    wb = Workbook()
    ws = wb.active
    ws.title = "All_Matches"

    green_fill = PatternFill(start_color="00C853", end_color="00C853", fill_type="solid")
    red_fill = PatternFill(start_color="D50000", end_color="D50000", fill_type="solid")

    block_width = 4  # Round | S1 | S2 | (blank)
    current_col = 1

    for match in match_data:
        history1 = match["history1"]
        history2 = match["history2"]
        player1 = match["player1"]
        player2 = match["player2"]

        rounds = len(history1)

        # Headers
        ws.cell(row=1, column=current_col).value = "Round"
        ws.cell(row=1, column=current_col + 1).value = player1
        ws.cell(row=1, column=current_col + 2).value = player2

        # Fill rows
        for i in range(rounds):
            row_index = i + 2

            ws.cell(row=row_index, column=current_col).value = i + 1

            cell1 = ws.cell(row=row_index, column=current_col + 1)
            cell2 = ws.cell(row=row_index, column=current_col + 2)

            cell1.value = ""
            cell2.value = ""

            cell1.fill = green_fill if history1[i] == "C" else red_fill
            cell2.fill = green_fill if history2[i] == "C" else red_fill

            cell1.alignment = Alignment(horizontal="center")
            cell2.alignment = Alignment(horizontal="center")

        # Set column widths
        ws.column_dimensions[ws.cell(row=1, column=current_col).column_letter].width = 8
        ws.column_dimensions[ws.cell(row=1, column=current_col + 1).column_letter].width = 4
        ws.column_dimensions[ws.cell(row=1, column=current_col + 2).column_letter].width = 4

        # Move to next block
        current_col += block_width

    filename = f"All_Matches_noise={NOISE}_rounds={ROUNDS}.xlsx"
    wb.save(filename)

    print(f"Exported all matches to {filename}")


if __name__ == "__main__":
    main()
